import cv2
import sys
from PyQt5.QtWidgets import  QCheckBox, QLineEdit, QPushButton, QSlider, QStatusBar, QTextEdit, QWidget, QLabel, QApplication,QMainWindow
from PyQt5.QtCore import QThread, Qt, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QImage, QPixmap
from PyQt5 import uic
import numpy as np
from pyModbusTCP.server import ModbusServer, DataBank
import streamVideoFlask
import time
import openpyxl



 

class enviaImagem(QThread):
    
    
    def __init__(self):
        super(enviaImagem, self).__init__()
 
    
    
    def run(self):
        
        
        streamVideoFlask.start()    

       
    
class Thread(QThread):

   

    changePixmap = pyqtSignal(QImage)
    changePixmap2 = pyqtSignal(QImage)
    contTotal=pyqtSignal(str)
       
  

    
    width = 150
    height = 400
    #eggCount = 0
    exitCounter = 0
    #OffsetRefLines = 100 # Adjust ths value according to your usage
    ReferenceFrame = None
    distance_tresh = 400
    #radius_min = 0
    #radius_max = 0
    #area_min = 750
    #area_max = 1100

 

    
 



    def __init__(self,modbusHab,areaMAx,areaMin,contorno,radiusMin, radiusMax, of7Linhas,portaTCP,porcEsc,contagemParou,diaAtual):
        super(Thread, self).__init__()
        self.area_max=areaMAx
        self.area_min=areaMin
        self.contornoValue=contorno
        self.radiusMin=radiusMin        
        self.radiusMax=radiusMax
        self.of7Linhas=of7Linhas
        self.eggCount=int(contagemParou)
        self.portaTCP=portaTCP
        self.modbusHab=modbusHab
        self.cont=0
        self.porcEsc=porcEsc
        self.diaAtual=diaAtual
        
      
        #self.modbus=TCP

            
        
    def stop(self):
        self.coreActive=False
        self.quit()
        
    def lerModbus(self):
        zera=DataBank.get_words(1)
        if int(zera[0])==1:
            self.eggCount=0
            zera[0]=0
            DataBank.set_words(1, [0])
        

    def escreveModbus(self,contador):        

            self.server = ModbusServer("127.0.0.1", int(self.portaTCP), no_block=True)
            self.server.start()                           
            DataBank.set_words(0, [int(contador)])

    def run(self):

        
        self.coreActive=True
        cap = cv2.VideoCapture('13.mp4')
        #cap = cv2.VideoCapture(0)
        fgbg = cv2.createBackgroundSubtractorMOG2()  # for mask

        tempoini=time.time()
        


        while (self.coreActive):
         
            data=time.localtime()
            dia=data.tm_mday
            if (int(dia)!=int(self.diaAtual)):
                self.eggCount=0


            (grabbed, frame) = cap.read()
            frame = cv2.rotate(frame,cv2.ROTATE_90_COUNTERCLOCKWISE)
           

      

            if not grabbed:
                print('Egg count: ' + str(self.eggCount))
                print('\n End of the video file...')
                break

            # get Settings radius/area values
            #radius_min=10
            #radius_max = 20
            #area_min=500
            #area_max = 1000
            borderSize =self.contornoValue
         

            if self.radiusMin == '':
                self.radiusMin = 0
            if self.radiusMax == '':
                self.radiusMax = 0

            if self.area_min == '':
                self.area_min = 0
            if self.area_max == '':
                self.area_max = 0

            
            percent=int(self.porcEsc)
            width = int(frame.shape[1] * percent // 100)
            height = int(frame.shape[0] * percent // 100)
            dim = (width, height)
            frame40=cv2.resize(frame, dim, interpolation=cv2.INTER_AREA)


            height = np.size(frame40, 0)
            width = np.size(frame40, 1)

            fgmask = fgbg.apply(frame40)

            hsv = cv2.cvtColor(frame40, cv2.COLOR_BGR2HSV)
            th, bw = cv2.threshold(hsv[:, :, 2], 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
            kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
            morph = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, kernel)
            dist = cv2.distanceTransform(morph, cv2.DIST_L2, cv2.DIST_MASK_PRECISE)

            distborder = cv2.copyMakeBorder(dist, borderSize, borderSize, borderSize, borderSize,
                                            cv2.BORDER_CONSTANT | cv2.BORDER_ISOLATED, 0)
            gap = 10
            kernel2 = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2 * (borderSize - gap) + 1, 2 * (borderSize - gap) + 1))
            kernel2 = cv2.copyMakeBorder(kernel2, gap, gap, gap, gap,
                                        cv2.BORDER_CONSTANT | cv2.BORDER_ISOLATED, 0)

            distTempl = cv2.distanceTransform(kernel2, cv2.DIST_L2, cv2.DIST_MASK_PRECISE)

            nxcor = cv2.matchTemplate(distborder, distTempl, cv2.TM_CCOEFF_NORMED)

            mn, mx, _, _ = cv2.minMaxLoc(nxcor)
            th, peaks = cv2.threshold(nxcor, mx * 0.5, 255, cv2.THRESH_BINARY)
            peaks8u = cv2.convertScaleAbs(peaks)

            # fgmask = self.fgbg.apply(peaks8u)

            #_, contours, hierarchy = cv2.findContours(peaks8u, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            contours, hierarchy = cv2.findContours(peaks8u, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
            peaks8u = cv2.convertScaleAbs(peaks)  # to use as mask

            # plot reference lines (entrance and exit lines)
            coordYEntranceLine = (height // 2) + int(self.of7Linhas)
            coordYMiddleLine = (height // 2)
            coordYExitLine = (height // 2) - int(self.of7Linhas)
            cv2.line(frame40, (0, coordYEntranceLine), (width, coordYEntranceLine), (255, 0, 0), 2)
            cv2.line(frame40, (0, coordYMiddleLine), (width, coordYMiddleLine), (0, 255, 0), 6)
            cv2.line(frame40, (0, coordYExitLine), (width, coordYExitLine), (255, 0, 0), 2)
  

            flag = False
            egg_list = []
            egg_index = 0

            for i in range(len(contours)):
                contour = contours[i]

                (x, y), radius = cv2.minEnclosingCircle(contour)
                radius = int(radius)

                (x, y, w, h) = cv2.boundingRect(contour)

                egg_index = i

                egg_list.append([x, y, flag])
                

                if len(contour) >= 5:

                    if (radius <= int(self.radiusMax) and radius >= int(self.radiusMin)):

                        # print("radius: ", radius)
                        # pprint.pprint(hierarchy)

                        ellipse = cv2.fitEllipse(contour)
                        # (x, y, w, h) = cv2.boundingRect(contour)
                    
                        (center, axis, angle) = ellipse
                        try:
                            coordXContour, coordYContour = int(center[0]), int(center[1])
                        except:
                            coordXContour, coordYContour=0,0
                        coordXCentroid = (2 * coordXContour + w) // 2
                        coordYCentroid = (2 * coordYContour + h) // 2
                     
                        try:
                            ax1, ax2 = int(axis[0]) - 2, int(axis[1]) - 2
                        except:
                            ax1,ax2=2,2
                           
                        orientation = float(angle)
                        area = cv2.contourArea(contour)

                        if area >= float(self.area_min) and area <= float(self.area_max):
                            #print('egg list: ' + str(egg_list) + ' index: ' + str(egg_index))

                            
                            if ((coordYContour <= coordYEntranceLine) and (coordYContour >= coordYExitLine)):

                            
                                cv2.ellipse(frame40, (coordXContour, coordYContour), (ax1, ax2), orientation, 0, 360,
                                            (255, 0, 0), 2)  # blue
                                cv2.circle(frame40, (coordXContour, coordYContour), 1, (0, 255, 0), 15)  # green
                                                            
                                cv2.putText(frame40, str(int(area)), (coordXContour, coordYContour), cv2.FONT_HERSHEY_SIMPLEX,
                                       0.5, 0, 1, cv2.LINE_AA)

                            for k in range(len(egg_list)):
                                  egg_new_X = x
                                  egg_new_Y = y
                         

                            dist = abs(egg_new_Y - egg_list[k][1])

                            if dist > self.distance_tresh:  # distance_tresh = 200
                                    egg_list.append([egg_new_X, egg_new_Y, flag])
                                    

                            absDistance = abs(coordYContour - coordYEntranceLine)

                            if ((coordYContour >= coordYEntranceLine) and (absDistance <= 3)):
                                                    
                                   self.eggCount += 1
                                   egg_list.remove([egg_new_X, egg_new_Y, flag])

            #salva imagem a cada 2 segundos
            tempofinal=time.time()
            if (tempofinal-tempoini>2):
                cv2.imwrite('frame.jpg',frame40)
                tempoini=time.time()
            #escreve e ler modbus a cada 2 segundos 
                if (self.modbusHab==1):
                    self.escreveModbus(self.eggCount)
                    self.lerModbus()
                                            
            self.rgbImage = cv2.cvtColor(frame40, cv2.COLOR_BGR2RGB)
            h, w, ch = self.rgbImage.shape
            bytesPerLine = ch * w
            convertToQtFormat = QImage(self.rgbImage.data, w, h, bytesPerLine, QImage.Format_RGB888)
            p = convertToQtFormat.scaled(640, 480, Qt.KeepAspectRatio)
            self.changePixmap.emit(p)

        
            rgbImage2 = cv2.cvtColor(peaks8u, cv2.COLOR_BGR2RGB)
            h, w, ch = rgbImage2.shape
            bytesPerLine = ch * w
            convertToQtFormat = QImage(rgbImage2.data, w, h, bytesPerLine, QImage.Format_RGB888)
            p = convertToQtFormat.scaled(640, 480, Qt.KeepAspectRatio)
            self.changePixmap2.emit(p)

            self.contTotal.emit(str(self.eggCount))
            
         
                       
        #  cleanup the camera and close any open windows
      
        cap.release()
        
    

          

class ui(QMainWindow):
    
    def __init__(self):
        super(ui,self).__init__()
       
       
        #janela ajustes
        
        self.contorno=ajustes.findChild(QSlider,"horizontalSlider")
        self.areaMax=ajustes.findChild(QLineEdit,"lineEdit")
        self.areaMin=ajustes.findChild(QLineEdit,"lineEdit2")
        self.radiusMin=ajustes.findChild(QLineEdit,"lineEdit2_2")
        self.radiusMax=ajustes.findChild(QLineEdit,"lineEdit_2")
        self.of7linhas=ajustes.findChild(QLineEdit,"lineEdit_3")
        self.botaoAtualizar=ajustes.findChild(QPushButton,"pushButton_3")
        self.portaTCP=ajustes.findChild(QLineEdit,"lineEdit2_3")
        self.modbusSim=ajustes.findChild(QCheckBox,"checkBox")
        self.modbusNao=ajustes.findChild(QCheckBox,"checkBox_2")
        self.porcEsc=ajustes.findChild(QLineEdit,"lineEdit2_5")

         #janela principal
        self.botaoIniciar=principal.findChild(QPushButton,"pushButton")
        self.botaoParar=principal.findChild(QPushButton,"pushButton_2")
        self.mostraimagem=principal.findChild(QLabel,"label")
        self.mostraimagem2=principal.findChild(QLabel,"label_2")
        self.qtdOvos=principal.findChild(QLabel,"label_4")
        self.botaoAjustes=principal.findChild(QPushButton,"pushButton_4")
        self.botaoZerar=principal.findChild(QPushButton,"pushButton_3")
        self.status=principal.findChild(QStatusBar,"statusbar")

        self.contorno.valueChanged.connect(self.valorContorno)
        self.botaoIniciar.clicked.connect(self.botaoIniciarFunc)
        self.botaoParar.clicked.connect(self.botaoPararFunc)
        self.botaoAtualizar.clicked.connect(self.botaoAtualizarFunc)
        self.botaoAjustes.clicked.connect(self.botaoAjustesFunc)
        self.modbusSim.toggled.connect(self.habilitaMod)
        self.modbusNao.toggled.connect(self.desabilitaMod)
        self.botaoZerar.clicked.connect(self.botaoZerarFunc)
        self.status.showMessage('Servidor online em: 192.168.1.11/1880                                                                            Developed by Jeison')

        
        
        principal.show()

        self.botaoIniciarFunc()
                
        


        
    def botaoAjustesFunc(self):
        ajustes.show()



    def botaoAtualizarFunc(self):
        self.th.stop()
        self.salvaPlanilha()
        self.botaoIniciarFunc()
       
 

    def botaoIniciarFunc(self):

        
        self.lerVariaveis()
        self.lerContagem()
        
        self.th = Thread(self.modbusSim.isChecked(),self.areaMax.text(), self.areaMin.text(),self.valorContorno(),self.radiusMin.text(),
        self.radiusMax.text(),self.of7linhas.text(),self.portaTCP.text(),self.porcEsc.text(),self.contagemParou,self.diaAtual)
        self.th.start()
        self.th.changePixmap.connect(self.setImage)
        self.th.changePixmap2.connect(self.setImage2)
        self.th.contTotal.connect(self.mostraTotalOvos)
       
    
        
        self.th1=enviaImagem()
        self.th1.start()
 

        self.botaoIniciar.setEnabled(False)
        self.botaoParar.setEnabled(True)

    def botaoZerarFunc(self):
        self.th.stop()
        self.ovos=0
        self.salvaContagem()
        self.botaoIniciarFunc()

    def botaoPararFunc(self):

        self.th.stop()
        self.botaoIniciar.setEnabled(True)
        self.botaoParar.setEnabled(False)
   
    def habilitaMod(self):
        self.modbusNao.setChecked(False)
        self.portaTCP.setEnabled(True)
        self.analise=1
      
  
    def desabilitaMod(self):
        self.analise=0
        self.modbusSim.setChecked(False)
        self.portaTCP.setEnabled(False)
        

      
    def valorContorno (self):
        return self.contorno.value()
        


    def mostraTotalOvos (self,valor1):
        valor=str(self.valorContorno())
        self.qtdOvos.setText(valor1)
        self.ovos=valor1
        self.salvaContagem()


    def salvaContagem(self):
        data=time.localtime()
        dia=int(data.tm_mday)
     
        
        planilha=openpyxl.Workbook()
        Dados = planilha['Sheet']
        Dados.title = 'Contagem'
        Dados['A1'] = 'Dia'
        Dados['B1'] = 'Valor'

        Dados.cell(column=1, row=2, value=dia)
        Dados.cell(column=2, row=2, value=self.ovos)
      
        planilha.save("Contagem.xlsx")

    def salvaPlanilha(self):

        planilha=openpyxl.Workbook()
        variaveis = planilha['Sheet']
        variaveis.title = 'Variáveis'
        variaveis['A1'] = 'Nome'
        variaveis['B1'] = 'Valor'

        variaveis.cell(column=1, row=2, value='areaMax')
        variaveis.cell(column=2, row=2, value=self.areaMax.text())
        
        variaveis.cell(column=1, row=3, value='areaMin')
        variaveis.cell(column=2, row=3, value=self.areaMin.text())

        variaveis.cell(column=1, row=4, value='Contorno')
        variaveis.cell(column=2, row=4, value=self.valorContorno())

        variaveis.cell(column=1, row=5, value='RadiusMin')
        variaveis.cell(column=2, row=5, value=self.radiusMin.text())

        variaveis.cell(column=1, row=6, value='RadiusMax')
        variaveis.cell(column=2, row=6, value=self.radiusMax.text())

        variaveis.cell(column=1, row=7, value='Of7Linhas')
        variaveis.cell(column=2, row=7, value=self.of7linhas.text())

        variaveis.cell(column=1, row=8, value='PortaTCP')
        variaveis.cell(column=2, row=8, value=self.portaTCP.text())

        variaveis.cell(column=1, row=9, value='PorcEscala')
        variaveis.cell(column=2, row=9, value=self.porcEsc.text())

        planilha.save("Variáveis salvas.xlsx")

    def lerVariaveis(self):
        wb = openpyxl.load_workbook("Variáveis salvas.xlsx", data_only=True)
        sh = wb["Variáveis"]
        self.areaMax.setText(str(sh["B2"].value))
        self.areaMin.setText(str(sh["B3"].value))
        self.contorno.setValue(int(sh["B4"].value))
        self.radiusMin.setText(str(sh["B5"].value))
        self.radiusMax.setText(str(sh["B6"].value))
        self.of7linhas.setText(str(sh["B7"].value))
        self.portaTCP.setText(str(sh["B8"].value))
        self.porcEsc.setText(str(sh["B9"].value))

    def lerContagem(self):
        wb = openpyxl.load_workbook("Contagem.xlsx", data_only=True)
        sh = wb["Contagem"]
        self.contagemParou=(str(sh["B2"].value))
        self.diaAtual=(str(sh["A2"].value))
      


    @pyqtSlot(QImage)
    def setImage(self, image):
        self.mostraimagem.setPixmap(QPixmap.fromImage(image))

    def setImage2(self, image):
        self.mostraimagem2.setPixmap(QPixmap.fromImage(image))
     
       
    
 
app = QApplication(sys.argv)
principal=uic.loadUi("OpenCv.ui")
ajustes=uic.loadUi("Config.ui")
uiWindow=ui()
app.exec()